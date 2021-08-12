import pathlib
import csv
import openpyxl
from openpyxl.styles import Protection
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import win32com.client as win32

#【0】 初期設定　データ開始行の入力
print('データ開始行は？？')
start_row = int(input('>>: '))

#【１】 エクセルファイル調査票の一括読み込み
path = pathlib.Path()
for pass_obj in path.iterdir():
    if pass_obj.match("*.xlsx"):
        #   エクセルファイルのオープン
        owb = openpyxl.load_workbook(pass_obj)
        sh = owb["点検修繕調査様式 (提出用)"]

        print(sh.max_row, sh.max_column, sh.cell(row = sh.max_row, column = sh.max_column).coordinate)

        #   シート保護の解除
        sh.protection.disable()

        #   ブック保護の解除
        owb.security = WorkbookProtection(workbookPassword='mlit8111', lockStructure=False)
        
        #   配布時テーブル設定を解除
        sh.ListObjects(1).Unlist

        #   任意の範囲に対してテーブル設定
        table = "A" + str(start_row - 1) + ":" + str(sh.cell(row = sh.max_row, column = sh.max_column).coordinate)
        tab = Table(displayName = "Table2", ref = table)
        style = TableStyleInfo(name= "TableStyleMedium15", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        sh.add_table(tab)
        
        #   ワークブックの保存
        owb.save(pass_obj)