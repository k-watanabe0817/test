import pathlib
import csv
import openpyxl
from openpyxl.styles import Protection
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill
from openpyxl.styles import numbers
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from tqdm import tqdm
import subprocess

#【0】 初期設定
#   新規施設の入力用データ数
NEW_data = 10

#   調査票書き込み開始行
start_row = 13

#【１】 雛型ファイル名、パスワード当の指定
print('調査様式雛型Excelを指定してください（拡張子含む）')
excel_path = input('>>: ')

print('調査様式のシート保護解除時パスワードを入力してください')
password = input('>>: ')

#【２】 BIで整理したCSVの複数読み込み,基礎自治体単位にデータを辞書に整理
#   PowerBIから出力したデータセットcsvの読み込み
path = pathlib.Path()
data = {}
for pass_obj in path.iterdir():
    if pass_obj.match("*.csv"):
        #   CSVファイルのオープン
        with open(pass_obj, encoding = "utf-8_sig") as f:
            #   CSV読み込み
            reader = csv.reader(f)

            #   ラベルデータの読み込みをスキップする
            header = next(reader)

            for read in reader:

                #   データを団体別に辞書に格納
                Prefectures_No = read[85]               #団体コード
                PLACE = read[0]                         #連番
                data.setdefault(Prefectures_No,{
                    "団体コード":Prefectures_No,
                    '事務所名':read[14]
                    }
                )
                data[Prefectures_No].setdefault(PLACE,{
                    '国交省作業用番号':read[0],
                    '地整番号':read[1],
                    '作業内容フラグ_区分':read[2],
                    '作業内容フラグ_内訳':read[3],
                    '諸元_施設名_種別':read[4],
                    '諸元_施設名_形式':read[5],
                    '諸元_施設名_路側式の確認':read[6],
                    '諸元_施設設置場所_択一式':read[7],
                    '諸元_施設設置場所_記述式':read[8],
                    '諸元_路線_路線名':read[9],
                    '諸元_路線_道路種別':read[10],
                    '諸元_管理者_区分':read[11],
                    '諸元_管理者_管理者名':read[12],
                    '諸元_管理者_事務所名':read[13],
                    '諸元_管理者_出張所名':read[14],
                    '諸元_設置年度（西暦）':read[15],
                    '諸元_管理番号':read[16],
                    '諸元_幅員（m）':read[17],
                    '諸元_国管理の標識柱に他の道路管理者等が標識等を添架している場合にその管理者を記入':read[18],
                    '諸元_占用物件（名称）':read[19],
                    '諸元_国以外の道路管理者等の標識柱に国が標識等を添架している場合にその柱の管理者を記入':read[20],
                    '諸元_行政区域_都道府県名':read[21],
                    '諸元_行政区域_市区町村名':read[22],
                    '諸元_位置_起点_緯度':read[23],
                    '諸元_位置_起点_経度':read[24],
                    '諸元_位置_終点_緯度':read[25],
                    '諸元_位置_終点_経度':read[26],
                    '諸元_距離標（キロポスト）':read[27],
                    '諸元_上り・下り線':read[28],
                    '諸元_緊急輸送道路':read[29],
                    '諸元_自専道・一般道':read[30],
                    '老朽化対策に関する調査項目_直近で実施した点検の点検要領':read[31],
                    '老朽化対策に関する調査項目_点検計画_2010':read[32],
                    '老朽化対策に関する調査項目_点検計画_2011':read[33],
                    '老朽化対策に関する調査項目_点検計画_2012':read[34],
                    '老朽化対策に関する調査項目_点検計画_2013':read[35],
                    '老朽化対策に関する調査項目_点検計画_2014':read[36],
                    '老朽化対策に関する調査項目_点検計画_2015':read[37],
                    '老朽化対策に関する調査項目_点検計画_2016':read[38],
                    '老朽化対策に関する調査項目_点検計画_2017':read[39],
                    '老朽化対策に関する調査項目_点検計画_2018':read[40],
                    '老朽化対策に関する調査項目_点検計画_2019':read[41],
                    '老朽化対策に関する調査項目_点検計画_2020':read[42],
                    '老朽化対策に関する調査項目_点検計画_2021':read[43],
                    '老朽化対策に関する調査項目_点検計画_2022':read[44],
                    '老朽化対策に関する調査項目_点検計画_2023':read[45],
                    '老朽化対策に関する調査項目_点検計画_備考':read[46],
                    '老朽化対策に関する調査項目_点検計画_定期点検（詳細、中間）実施年度':read[47],
                    '老朽化対策に関する調査項目_点検計画_定期点検（詳細、中間）実施年月日':read[48],
                    '老朽化対策に関する調査項目_点検記録_対策の要否_支柱':read[49],
                    '老朽化対策に関する調査項目_点検記録_対策の要否_横梁':read[50],
                    '老朽化対策に関する調査項目_点検記録_対策の要否_標識板等':read[51],
                    '老朽化対策に関する調査項目_点検記録_対策の要否_基礎':read[52],
                    '老朽化対策に関する調査項目_点検記録_対策の要否_その他':read[53],
                    '老朽化対策に関する調査項目_点検記録_対策の要否_施設単位':read[54],
                    '老朽化対策に関する調査項目_点検記録_判定区分_H22要領に基づく損傷度判定':read[55],
                    '老朽化対策に関する調査項目_点検記録_判定区分_H26・H31要領に基づく損傷度判定':read[56],
                    '老朽化対策に関する調査項目_点検記録_判定区分_H26・H31要領に基づく健全性判定':read[57],
                    '老朽化対策に関する調査項目_点検記録_点検表記録様式ファイル名':read[58],
                    '老朽化対策に関する調査項目_点検記録_合いマーク実施の有無':read[59],
                    '老朽化対策に関する調査項目_点検記録_落下防止ワイヤーの設置の有無':read[60],
                    '老朽化対策に関する調査項目_点検記録_所見等（任意）':read[61],
                    '老朽化対策に関する調査項目_修繕計画（任意）_2010':read[62],
                    '老朽化対策に関する調査項目_修繕計画（任意）_2011':read[63],
                    '老朽化対策に関する調査項目_修繕計画（任意）_2012':read[64],
                    '老朽化対策に関する調査項目_修繕計画（任意）_2013':read[65],
                    '老朽化対策に関する調査項目_修繕計画（任意）_2014':read[66],
                    '老朽化対策に関する調査項目_修繕計画（任意）_2015':read[67],
                    '老朽化対策に関する調査項目_修繕計画（任意）_2016':read[68],
                    '老朽化対策に関する調査項目_修繕計画（任意）_2017':read[69],
                    '老朽化対策に関する調査項目_修繕計画（任意）_2018':read[70],
                    '老朽化対策に関する調査項目_修繕計画（任意）_2019':read[71],
                    '老朽化対策に関する調査項目_修繕計画（任意）_2020':read[72],
                    '老朽化対策に関する調査項目_修繕計画（任意）_2021':read[73],
                    '老朽化対策に関する調査項目_修繕計画（任意）_2022':read[74],
                    '老朽化対策に関する調査項目_修繕計画（任意）_2023':read[75],
                    '老朽化対策に関する調査項目_修繕内容':read[76],
                    '老朽化対策に関する調査項目_概算修繕金額（百万円）':read[77],
                    '老朽化対策に関する調査項目_修繕設計_着手済':read[78],
                    '老朽化対策に関する調査項目_修繕設計_完了年月':read[79],
                    '老朽化対策に関する調査項目_修繕_着手済':read[80],
                    '老朽化対策に関する調査項目_修繕_完了年月':read[81],
                    '老朽化対策に関する調査項目_措置記録_再判定実施年月日':read[82],
                    '老朽化対策に関する調査項目_措置記録_再判定区分':read[83],
                    '老朽化対策に関する調査項目_備考':read[84]
                    }
                )

#   プログレスバーの設定
bar = tqdm(total = len(data))
bar.set_description('Progress rate')

#【３】 雛型ファイルから各団体に配布する調査様式を作成
#   基礎自治体単位で調査様式の作成
for Prefectures_No in data.values():
    #   雛型ファイルを開く
    owb = openpyxl.load_workbook(excel_path)
    sh = owb.active

    #   施設種別に調査様式に書き込み
    row = start_row                 #   書き込み行情報初期化
    data_count = len(Prefectures_No.values())
    
    for PLACE_data in Prefectures_No.values():  
        if isinstance(PLACE_data, dict):
            #   調査票に書き込み
            sh.cell(row,1).value = PLACE_data['国交省作業用番号']
            sh.cell(row,2).value = PLACE_data['地整番号']
            sh.cell(row,3).value = PLACE_data['作業内容フラグ_区分']
            sh.cell(row,4).value = PLACE_data['作業内容フラグ_内訳']
            sh.cell(row,5).value = PLACE_data['諸元_施設名_種別']
            sh.cell(row,6).value = PLACE_data['諸元_施設名_形式']
            sh.cell(row,7).value = PLACE_data['諸元_施設名_路側式の確認']
            sh.cell(row,8).value = PLACE_data['諸元_施設設置場所_択一式']
            sh.cell(row,9).value = PLACE_data['諸元_施設設置場所_記述式']
            sh.cell(row,10).value = PLACE_data['諸元_路線_路線名']
            sh.cell(row,11).value = PLACE_data['諸元_路線_道路種別']
            sh.cell(row,12).value = PLACE_data['諸元_管理者_区分']
            sh.cell(row,13).value = PLACE_data['諸元_管理者_管理者名']
            sh.cell(row,14).value = PLACE_data['諸元_管理者_事務所名']
            sh.cell(row,15).value = PLACE_data['諸元_管理者_出張所名']
            sh.cell(row,16).value = PLACE_data['諸元_設置年度（西暦）']
            sh.cell(row,17).value = PLACE_data['諸元_管理番号']
            sh.cell(row,18).value = PLACE_data['諸元_幅員（m）']
            sh.cell(row,19).value = PLACE_data['諸元_国管理の標識柱に他の道路管理者等が標識等を添架している場合にその管理者を記入']
            sh.cell(row,20).value = PLACE_data['諸元_占用物件（名称）']
            sh.cell(row,21).value = PLACE_data['諸元_国以外の道路管理者等の標識柱に国が標識等を添架している場合にその柱の管理者を記入']
            sh.cell(row,22).value = PLACE_data['諸元_行政区域_都道府県名']
            sh.cell(row,23).value = PLACE_data['諸元_行政区域_市区町村名']
            sh.cell(row,24).value = PLACE_data['諸元_位置_起点_緯度']
            sh.cell(row,25).value = PLACE_data['諸元_位置_起点_経度']
            sh.cell(row,26).value = PLACE_data['諸元_位置_終点_緯度']
            sh.cell(row,27).value = PLACE_data['諸元_位置_終点_経度']
            sh.cell(row,28).value = PLACE_data['諸元_距離標（キロポスト）']
            sh.cell(row,29).value = PLACE_data['諸元_上り・下り線']
            sh.cell(row,30).value = PLACE_data['諸元_緊急輸送道路']
            sh.cell(row,31).value = PLACE_data['諸元_自専道・一般道']
            sh.cell(row,32).value = PLACE_data['老朽化対策に関する調査項目_直近で実施した点検の点検要領']
            sh.cell(row,33).value = PLACE_data['老朽化対策に関する調査項目_点検計画_2010']
            sh.cell(row,34).value = PLACE_data['老朽化対策に関する調査項目_点検計画_2011']
            sh.cell(row,35).value = PLACE_data['老朽化対策に関する調査項目_点検計画_2012']
            sh.cell(row,36).value = PLACE_data['老朽化対策に関する調査項目_点検計画_2013']
            sh.cell(row,37).value = PLACE_data['老朽化対策に関する調査項目_点検計画_2014']
            sh.cell(row,38).value = PLACE_data['老朽化対策に関する調査項目_点検計画_2015']
            sh.cell(row,39).value = PLACE_data['老朽化対策に関する調査項目_点検計画_2016']
            sh.cell(row,40).value = PLACE_data['老朽化対策に関する調査項目_点検計画_2017']
            sh.cell(row,41).value = PLACE_data['老朽化対策に関する調査項目_点検計画_2018']
            sh.cell(row,42).value = PLACE_data['老朽化対策に関する調査項目_点検計画_2019']
            sh.cell(row,43).value = PLACE_data['老朽化対策に関する調査項目_点検計画_2020']
            sh.cell(row,44).value = PLACE_data['老朽化対策に関する調査項目_点検計画_2021']
            sh.cell(row,45).value = PLACE_data['老朽化対策に関する調査項目_点検計画_2022']
            sh.cell(row,46).value = PLACE_data['老朽化対策に関する調査項目_点検計画_2023']
            sh.cell(row,47).value = PLACE_data['老朽化対策に関する調査項目_点検計画_備考']
            sh.cell(row,48).value = PLACE_data['老朽化対策に関する調査項目_点検計画_定期点検（詳細、中間）実施年度']
            sh.cell(row,49).value = PLACE_data['老朽化対策に関する調査項目_点検計画_定期点検（詳細、中間）実施年月日']
            sh.cell(row,50).value = PLACE_data['老朽化対策に関する調査項目_点検記録_対策の要否_支柱']
            sh.cell(row,51).value = PLACE_data['老朽化対策に関する調査項目_点検記録_対策の要否_横梁']
            sh.cell(row,52).value = PLACE_data['老朽化対策に関する調査項目_点検記録_対策の要否_標識板等']
            sh.cell(row,53).value = PLACE_data['老朽化対策に関する調査項目_点検記録_対策の要否_基礎']
            sh.cell(row,54).value = PLACE_data['老朽化対策に関する調査項目_点検記録_対策の要否_その他']
            sh.cell(row,55).value = PLACE_data['老朽化対策に関する調査項目_点検記録_対策の要否_施設単位']
            sh.cell(row,56).value = PLACE_data['老朽化対策に関する調査項目_点検記録_判定区分_H22要領に基づく損傷度判定']
            sh.cell(row,57).value = PLACE_data['老朽化対策に関する調査項目_点検記録_判定区分_H26・H31要領に基づく損傷度判定']
            sh.cell(row,58).value = PLACE_data['老朽化対策に関する調査項目_点検記録_判定区分_H26・H31要領に基づく健全性判定']
            sh.cell(row,59).value = PLACE_data['老朽化対策に関する調査項目_点検記録_点検表記録様式ファイル名']
            sh.cell(row,60).value = PLACE_data['老朽化対策に関する調査項目_点検記録_合いマーク実施の有無']
            sh.cell(row,61).value = PLACE_data['老朽化対策に関する調査項目_点検記録_落下防止ワイヤーの設置の有無']
            sh.cell(row,62).value = PLACE_data['老朽化対策に関する調査項目_点検記録_所見等（任意）']
            sh.cell(row,63).value = PLACE_data['老朽化対策に関する調査項目_修繕計画（任意）_2010']
            sh.cell(row,64).value = PLACE_data['老朽化対策に関する調査項目_修繕計画（任意）_2011']
            sh.cell(row,65).value = PLACE_data['老朽化対策に関する調査項目_修繕計画（任意）_2012']
            sh.cell(row,66).value = PLACE_data['老朽化対策に関する調査項目_修繕計画（任意）_2013']
            sh.cell(row,67).value = PLACE_data['老朽化対策に関する調査項目_修繕計画（任意）_2014']
            sh.cell(row,68).value = PLACE_data['老朽化対策に関する調査項目_修繕計画（任意）_2015']
            sh.cell(row,69).value = PLACE_data['老朽化対策に関する調査項目_修繕計画（任意）_2016']
            sh.cell(row,70).value = PLACE_data['老朽化対策に関する調査項目_修繕計画（任意）_2017']
            sh.cell(row,71).value = PLACE_data['老朽化対策に関する調査項目_修繕計画（任意）_2018']
            sh.cell(row,72).value = PLACE_data['老朽化対策に関する調査項目_修繕計画（任意）_2019']
            sh.cell(row,73).value = PLACE_data['老朽化対策に関する調査項目_修繕計画（任意）_2020']
            sh.cell(row,74).value = PLACE_data['老朽化対策に関する調査項目_修繕計画（任意）_2021']
            sh.cell(row,75).value = PLACE_data['老朽化対策に関する調査項目_修繕計画（任意）_2022']
            sh.cell(row,76).value = PLACE_data['老朽化対策に関する調査項目_修繕計画（任意）_2023']
            sh.cell(row,77).value = PLACE_data['老朽化対策に関する調査項目_修繕内容']
            sh.cell(row,78).value = PLACE_data['老朽化対策に関する調査項目_概算修繕金額（百万円）']
            sh.cell(row,79).value = PLACE_data['老朽化対策に関する調査項目_修繕設計_着手済']
            sh.cell(row,80).value = PLACE_data['老朽化対策に関する調査項目_修繕設計_完了年月']
            sh.cell(row,81).value = PLACE_data['老朽化対策に関する調査項目_修繕_着手済']
            sh.cell(row,82).value = PLACE_data['老朽化対策に関する調査項目_修繕_完了年月']
            sh.cell(row,83).value = PLACE_data['老朽化対策に関する調査項目_措置記録_再判定実施年月日']
            sh.cell(row,84).value = PLACE_data['老朽化対策に関する調査項目_措置記録_再判定区分']
            sh.cell(row,85).value = PLACE_data['老朽化対策に関する調査項目_備考']
            row += 1
    
    #   記入箇所のセルロック、色塗り、文字列→数値に変換
    fill = PatternFill(patternType='solid', fgColor='FFFF99')
    for row in range(start_row, start_row + data_count + NEW_data):
        #   セルロック、色塗り
        for col in range(3, 85):
            sh.cell(row, col).fill = fill
            sh.cell(row, col).protection = Protection(locked=False)
        
        #   文字列→数字に変換
        
        for col in [16, 28]:
            if type(sh.cell(row, col).value) == str:
                try:
                    sh.cell(row, col).value = float(sh.cell(row, col).value)
                except ValueError:
                    continue

        #  縮小して全体を表示
        for col in range(3, 17):
            sh.cell(row, col).alignment = Alignment(shrinkToFit = "SHRINK_TO_FIT")
            
    
    #   任意の範囲に対してテーブル設定
    table = "A" + str(start_row - 1) + ":" + str(sh.cell(row = start_row + data_count + NEW_data - 1, column = sh.max_column).coordinate)
    tab = Table(displayName = "Table1", ref = table)
    style = TableStyleInfo(name= "TableStyleMedium15", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    sh.add_table(tab)

    #   このシートのすべてのユーザーに許可する操作
    sh.protection.objects = True                # オブジェクトの編集
    sh.protection.scenarios = True              # シナリオの編集
    sh.protection.formatCells = True            # セルの書式設定
    sh.protection.formatColumns = True          # 列の書式設定
    sh.protection.formatRows = True             # 行の書式設定
    sh.protection.insertColumns = True          # 列の挿入
    sh.protection.insertRows = True             # 行の挿入
    sh.protection.insertHyperlinks = True       # ハイパーリンクの挿入
    sh.protection.deleteColumns = True          # 列の削除
    sh.protection.deleteRows = True             # 行の削除
    sh.protection.selectLockedCells = True      # ロックされたセルの選択
    sh.protection.selectUnlockedCells = False   # ロックされていないセルの選択
    sh.protection.sort = True                   # 並べ替え
    sh.protection.autoFilter = False            # フィルター
    sh.protection.pivotTables = True            # ピボットテーブルレポート

    # パスワードをセット
    sh.protection.password = password

    #   シートの保護
    sh.protection.enable()

    #   ワークブックの保存
    owb.save('22小規模附属物-01道路標識_' + Prefectures_No['団体コード'] + Prefectures_No['事務所名'] + '.xlsx')
    bar.update(1)