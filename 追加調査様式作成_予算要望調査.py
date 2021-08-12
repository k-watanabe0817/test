import pathlib
import csv
import openpyxl
from openpyxl.styles import Protection
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill
from openpyxl.styles import numbers
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from tqdm import tqdm
import subprocess
import re

#【0】 初期設定
#   新規施設の入力用データ
NEW_data = 10

#   関数定義
def trans_word(inputtext, inputtable):
    return re.sub('({})'.format('|'.join(map(re.escape, inputtable.keys()))), lambda m: inputtable[m.group()], inputtext)

#   道路種別の読み替え対応表
table = {
    "高速自動車国道":"補助対象外",
    "一般国道（指定区間）":"補助対象外",
    "一般国道（指定区間外）":"補助国道",
    }

#【１】 雛型ファイル名、パスワード当の指定
print('調査様式雛型Excelを指定してください（拡張子含む）')
excel_path = input('>>: ')

print('調査様式のシート保護解除時パスワードを入力してください')
password = input('>>: ')

#【２】 BIで整理したCSVの複数読み込み,基礎自治体単位にデータを辞書に整理
#   PowerBIから出力したデータセットcsvの読み込み
path = pathlib.Path()
data = {}
for pass_obj in path.iterdir():
    if pass_obj.match("*.csv"):
        #   CSVファイルのオープン
        with open(pass_obj, encoding = "utf-8_sig") as f:
            #   CSV読み込み
            reader = csv.reader(f)

            #   ラベルデータの読み込みをスキップする
            header = next(reader)

            for read in reader:
                #   道路種別の読み替え
                if type(read[8]) == str:
                    try:
                        read[8] = trans_word(read[8], table)
                    except ValueError:
                        continue

                #   データを団体別、施設種別に辞書に格納
                TYPE = read[0]                          #施設種別
                Ppefectures_No = read[2] + read[4]      #団体コード+事業主体名
                data.setdefault(Ppefectures_No,{"団体コード":Ppefectures_No})
                data[Ppefectures_No].setdefault(TYPE,{"施設種別":TYPE})
                data[Ppefectures_No][TYPE].setdefault(read[1],{
                    "国交省作業用番号":read[1],
                    "団体コード":read[2],
                    "都道府県名":read[3],
                    "事業主体名":read[4],
                    "事業主体種別":read[5],
                    "施設名":read[6],
                    "路線_路線名":read[7],
                    "路線_道路種別":read[8],
                    "架設年度":read[9],
                    "施設延長":read[10],
                    "幅員":read[11],
                    "行政区域_市区町村":read[12],
                    "当該施設の緊急輸送道路種別":read[13],
                    "道路橋下状況_道路_緊急輸送道路種別":read[14],
                    "道路橋下状況_道路_道路橋下の管理者":read[15],
                    "道路橋下状況_鉄道":read[16],
                    "直近における点検結果_点検年度（西暦）":read[17],
                    "直近における点検結果_判定区分":read[18],
                    "1巡目点検_判定区分":read[19],
                    "全体事業費（百万円）":read[20],
                    "事業種別":read[21],
                    "事業着手年度":read[22],
                    "事業完了予定年度":read[23],
                    "事業内容_2021年度":read[24],
                    "新技術活用状況_活用有無":read[25],
                    "国費率":read[26],
                    "R３要望額_事業費":read[27]
                    }
                )

#   プログレスバーの設定
bar = tqdm(total = len(data))
bar.set_description('Progress rate')

#【３】 雛型ファイルから各団体に配布する調査様式を作成
#   基礎自治体単位で調査様式の作成
for Ppefectures_No in data.values():
    #   雛型ファイルを開く
    owb = openpyxl.load_workbook(excel_path)
    sh = owb.active

    #   施設種別に調査様式に書き込み
    row = 12     #   書き込み行情報初期化
    data_count = 0     #   カウント変数の初期化
    for TYPE_data in Ppefectures_No.values():
        if isinstance(TYPE_data, dict):
            data_count += len(TYPE_data) - 1
            for place_data in TYPE_data.values():
                if isinstance(place_data, dict):
                    #   調査票に書き込み
                    sh.cell(row, 1).value = place_data["団体コード"]
                    sh.cell(row, 2).value = place_data["都道府県名"]
                    sh.cell(row, 3).value = place_data["事業主体名"]
                    sh.cell(row, 4).value = place_data["事業主体種別"]
                    sh.cell(row, 5).value = place_data["国交省作業用番号"]
                    sh.cell(row, 6).value = place_data["施設名"]
                    sh.cell(row, 7).value = place_data["路線_路線名"]
                    sh.cell(row, 8).value = place_data["路線_道路種別"]                    
                    sh.cell(row, 10).value = place_data["架設年度"]
                    sh.cell(row, 11).value = place_data["施設延長"]
                    sh.cell(row, 12).value = place_data["幅員"]
                    sh.cell(row, 13).value = place_data["都道府県名"]
                    sh.cell(row, 14).value = place_data["行政区域_市区町村"]
                    sh.cell(row, 15).value = place_data["当該施設の緊急輸送道路種別"]
                    sh.cell(row, 16).value = place_data["道路橋下状況_道路_緊急輸送道路種別"]
                    sh.cell(row, 17).value = place_data["道路橋下状況_道路_道路橋下の管理者"]
                    sh.cell(row, 18).value = place_data["道路橋下状況_鉄道"]
                    sh.cell(row, 20).value = place_data["直近における点検結果_点検年度（西暦）"]
                    sh.cell(row, 21).value = place_data["直近における点検結果_判定区分"]
                    sh.cell(row, 22).value = place_data["1巡目点検_判定区分"]
                    sh.cell(row, 23).value = place_data["全体事業費（百万円）"]
                    sh.cell(row, 24).value = place_data["事業種別"]
                    sh.cell(row, 25).value = place_data["事業着手年度"]
                    sh.cell(row, 26).value = place_data["事業完了予定年度"]
                    sh.cell(row, 29).value = place_data["事業内容_2021年度"]
                    sh.cell(row, 34).value = place_data["新技術活用状況_活用有無"]
                    sh.cell(row, 52).value = place_data["国費率"]
                    sh.cell(row, 53).value = place_data["R３要望額_事業費"]
                    sh.cell(row, 60).value = place_data["国交省作業用番号"]
                    row += 1
    
    #   記入箇所のセルロック、色塗り、文字列→数値に変換
    fill = PatternFill(patternType='solid', fgColor='FFFF99')
    for row in range(12, data_count + 12 + NEW_data):
        #   セルロック、色塗り
        for col in [5, 6, 7, 8, 9, 10, 11, 12, 15, 16, 17, 18, 19, 20, 21, 24, 25, 26, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 55, 56, 57]:
            sh.cell(row, col).fill = fill
            sh.cell(row, col).protection = Protection(locked=False)

        #   文字列→数字に変換
        for col in [10, 11, 12, 20, 23, 25, 26, 52, 53]:
            if type(sh.cell(row, col).value) == str:
                try:
                    sh.cell(row, col).value = float(sh.cell(row, col).value)
                except ValueError:
                    continue

        #  縮小して全体を表示
        for col in [6, 7, 8, 14, 17, 18, 35, 36, 37, 38, 43, 44, 45]:
            sh.cell(row, col).alignment = Alignment(shrinkToFit = "SHRINK_TO_FIT")

        #   小数点桁数1桁表示設定
        for col in [11, 12, 40, 42]:
            sh.cell(row, col).number_format = '0.0'
        
        #   小数点桁数2桁表示設定
        sh.cell(row, 23).number_format = '0.000'
        for col in range(46, 57):
            sh.cell(row, col).number_format = '0.000'
       
        #   23列目に数式(全体事業費の計算)を入力
        sh.cell(row, 23).value ='=SUM(AT' + str(row) + ':AY'+ str(row) + ')+BA' + str(row) + '+SUM(BC' + str(row) + ':BE' + str(row) + ')'

        #   27列目に数式(新規・継続の判別)を入力
        sh.cell(row, 27).value = '=IF(Y' + str(row) + '="","-",IF(Y' + str(row) + '>=2021,"新規",IF(Y' + str(row) + '<=2020,"継続","")))'
        
        #   54列目に数式(国費率×R3事業費)を入力
        sh.cell(row, 54).value = '=AZ' + str(row) + '*BA' + str(row)
        
        #   58列目に数式(施設種別の判別)を入力
        sh.cell(row, 58).value = '=IF(LEFT(E' + str(row) + ',2)="BR","橋梁",IF(LEFT(E' + str(row) + ',2)="TU","トンネル",IF(LEFT(E' + str(row) + ',2)="CL","大型カルバート",IF(LEFT(E' + str(row) + ',2)="SH","シェッド",IF(LEFT(E' + str(row) + ',3)="FB1","横断歩道橋（跨線橋以外）",IF(LEFT(E' + str(row) + ',3)="FB2","横断歩道橋（跨線橋）",IF(LEFT(E' + str(row) + ',2)="GM","門型標識等",)))))))'
    
    #   新規施設用データ入力,セルロック、色塗り
    for row in range(data_count + 12, data_count + 12 + NEW_data):
        #新規施設用データ入力
        sh.cell(row, 1).value = sh.cell(12, 1).value
        sh.cell(row, 2).value = sh.cell(12, 2).value
        sh.cell(row, 3).value = sh.cell(12, 3).value
        sh.cell(row, 4).value = sh.cell(12, 4).value
        sh.cell(row, 60).value = '=E' + str(row)

        #   セルロック、色塗り
        sh.cell(row, 13).fill = fill
        sh.cell(row, 13).protection = Protection(locked=False)
        sh.cell(row, 14).fill = fill
        sh.cell(row, 14).protection = Protection(locked=False)

    #   任意の範囲に対してテーブル設定 ※今回はA5:BOnnセル
    table = "A11:BH" + str(data_count + 11 + NEW_data)
    tab = Table(displayName = "Table1", ref = table)
    style = TableStyleInfo(name= "TableStyleMedium15", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    sh.add_table(tab)

    #   このシートのすべてのユーザーに許可する操作
    sh.protection.objects = True                # オブジェクトの編集
    sh.protection.scenarios = True              # シナリオの編集
    sh.protection.formatCells = True            # セルの書式設定
    sh.protection.formatColumns = True          # 列の書式設定
    sh.protection.formatRows = True             # 行の書式設定
    sh.protection.insertColumns = True          # 列の挿入
    sh.protection.insertRows = True             # 行の挿入
    sh.protection.insertHyperlinks = True       # ハイパーリンクの挿入
    sh.protection.deleteColumns = True          # 列の削除
    sh.protection.deleteRows = True             # 行の削除
    sh.protection.selectLockedCells = True      # ロックされたセルの選択
    sh.protection.selectUnlockedCells = False   # ロックされていないセルの選択
    sh.protection.sort = True                   # 並べ替え
    sh.protection.autoFilter = False            # フィルター
    sh.protection.pivotTables = True            # ピボットテーブルレポート
    
    # パスワードをセット
    sh.protection.password = password

    #   シートの保護
    sh.protection.enable()

    # 他のシートを保護
    for sheet_name in['入力者チェック用シート＜措置状況＞', '入力者チェック用シート＜予算状況＞']:
        # 対象のワークシートオブジェクトを取得する
        ws = owb[sheet_name]

        #   シート（入力者チェック用シート＜措置状況＞）のすべてのユーザーに許可する操作
        ws.protection.objects = True                # オブジェクトの編集
        ws.protection.scenarios = True              # シナリオの編集
        ws.protection.formatCells = True            # セルの書式設定
        ws.protection.formatColumns = True          # 列の書式設定
        ws.protection.formatRows = True             # 行の書式設定
        ws.protection.insertColumns = True          # 列の挿入
        ws.protection.insertRows = True             # 行の挿入
        ws.protection.insertHyperlinks = True       # ハイパーリンクの挿入
        ws.protection.deleteColumns = True          # 列の削除
        ws.protection.deleteRows = True             # 行の削除
        ws.protection.selectLockedCells = True      # ロックされたセルの選択
        ws.protection.selectUnlockedCells = True    # ロックされていないセルの選択
        ws.protection.sort = True                   # 並べ替え
        ws.protection.autoFilter = True             # フィルター
        ws.protection.pivotTables = True            # ピボットテーブルレポート
        
        # パスワードをセット
        ws.protection.password = password

        #   シートの保護
        ws.protection.enable()

    #   ワークブックの保存
    owb.save(Ppefectures_No["団体コード"] + ".xlsx")
    bar.update(1)